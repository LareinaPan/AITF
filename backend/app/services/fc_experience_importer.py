from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

TEMPLATE_HEADERS = (
    "用例编号",
    "功能模块",
    "用例标题",
    "前置条件",
    "测试步骤",
    "预期结果",
    "优先级",
    "用例类型",
)

CASE_TYPE_ALIASES: dict[str, str] = {
    "positive": "positive",
    "正向": "positive",
    "negative": "negative",
    "异常": "negative",
    "boundary": "boundary",
    "边界": "boundary",
    "permission": "permission",
    "权限": "permission",
    "security": "security",
    "安全": "security",
    "compatibility": "compatibility",
    "兼容": "compatibility",
}

VALID_PRIORITIES = frozenset({"P0", "P1", "P2", "P3"})


class FcExperienceImportError(ValueError):
    """Raised when experience case Excel cannot be imported."""


@dataclass(frozen=True)
class ParsedExperienceCaseRow:
    row_number: int
    case_no: str | None
    module: str
    title: str
    preconditions: str | None
    steps: str
    expected_result: str
    priority: str
    case_type: str


@dataclass(frozen=True)
class ExperienceImportResult:
    valid_rows: list[ParsedExperienceCaseRow]
    errors: list[str]


def _cell_value(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_case_type(raw: str | None) -> str:
    if raw is None:
        return "positive"
    text = raw.strip()
    normalized = CASE_TYPE_ALIASES.get(text) or CASE_TYPE_ALIASES.get(text.lower())
    if normalized is None:
        raise FcExperienceImportError(f"不支持的用例类型：{raw}")
    return normalized


def normalize_priority(raw: str | None) -> str:
    if raw is None:
        return "P2"
    upper = raw.strip().upper()
    if upper not in VALID_PRIORITIES:
        raise FcExperienceImportError(f"不支持的优先级：{raw}")
    return upper


def parse_experience_case_excel(content: bytes) -> ExperienceImportResult:
    """Parse experience case rows from an Excel workbook."""
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise FcExperienceImportError(f"无法读取 Excel 文件：{exc}") from exc

    sheet = workbook.active
    if sheet is None:
        raise FcExperienceImportError("Excel 文件中没有工作表")

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise FcExperienceImportError("Excel 文件为空")

    header_row = [_cell_value(cell) or "" for cell in rows[0]]
    if header_row[: len(TEMPLATE_HEADERS)] != list(TEMPLATE_HEADERS):
        raise FcExperienceImportError(
            "Excel 表头不符合模板，请使用标准模板（用例编号、功能模块、用例标题…）"
        )

    valid_rows: list[ParsedExperienceCaseRow] = []
    errors: list[str] = []

    for index, row in enumerate(rows[1:], start=2):
        cells = list(row) + [None] * (len(TEMPLATE_HEADERS) - len(row))
        values = [_cell_value(cell) for cell in cells[: len(TEMPLATE_HEADERS)]]

        if all(value is None for value in values):
            continue

        case_no, module, title, preconditions, steps, expected_result, priority_raw, case_type_raw = values

        if not module or not title or not steps or not expected_result:
            errors.append(f"第 {index} 行缺少必填字段（功能模块/用例标题/测试步骤/预期结果）")
            continue

        try:
            priority = normalize_priority(priority_raw)
            case_type = normalize_case_type(case_type_raw)
        except FcExperienceImportError as exc:
            errors.append(f"第 {index} 行：{exc}")
            continue

        valid_rows.append(
            ParsedExperienceCaseRow(
                row_number=index,
                case_no=case_no,
                module=module,
                title=title,
                preconditions=preconditions,
                steps=steps,
                expected_result=expected_result,
                priority=priority,
                case_type=case_type,
            )
        )

    workbook.close()
    return ExperienceImportResult(valid_rows=valid_rows, errors=errors)
