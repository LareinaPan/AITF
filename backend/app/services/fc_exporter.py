import json
import re
import uuid
import zipfile
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.services.fc_experience_importer import TEMPLATE_HEADERS

CASE_TYPE_EXPORT_LABELS: dict[str, str] = {
    "positive": "正向",
    "negative": "异常",
    "boundary": "边界",
    "permission": "权限",
    "security": "安全",
    "compatibility": "兼容",
}


class FcExportError(ValueError):
    """Raised when functional test cases cannot be exported."""


@dataclass(frozen=True)
class ExportableFcCase:
    case_no: str
    module: str
    title: str
    preconditions: str | None
    steps: str
    expected_result: str
    priority: str
    case_type: str


def case_type_export_label(case_type: str) -> str:
    return CASE_TYPE_EXPORT_LABELS.get(case_type, case_type)


def safe_export_filename(name: str, suffix: str) -> str:
    """Build an ASCII-only filename safe for HTTP Content-Disposition headers."""
    stem = re.sub(r"[^A-Za-z0-9\-_]+", "_", name.strip()).strip("_") or "fc-export"
    return f"{stem[:64]}{suffix}"


def case_to_excel_row(case: ExportableFcCase) -> tuple[str, ...]:
    return (
        case.case_no,
        case.module,
        case.title,
        case.preconditions or "",
        case.steps,
        case.expected_result,
        case.priority,
        case_type_export_label(case.case_type),
    )


def export_cases_to_excel(cases: list[ExportableFcCase], *, project_name: str) -> bytes:
    """Export functional test cases to Excel bytes using appendix A column order."""
    if not cases:
        raise FcExportError("No test cases to export")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "功能用例"
    sheet.append(list(TEMPLATE_HEADERS))

    header_font = Font(bold=True)
    for cell in sheet[1]:
        cell.font = header_font

    for case in cases:
        sheet.append(list(case_to_excel_row(case)))

    sheet.freeze_panes = "A2"

    for index, _header in enumerate(TEMPLATE_HEADERS, start=1):
        column = get_column_letter(index)
        sheet.column_dimensions[column].width = 18 if index <= 3 else 24

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _new_xmind_topic(title: str, children: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    topic: dict[str, Any] = {
        "id": uuid.uuid4().hex,
        "class": "topic",
        "title": title,
    }
    if children:
        topic["children"] = {"attached": children}
    return topic


def _build_case_topic(case: ExportableFcCase) -> dict[str, Any]:
    attributes = f"优先级: {case.priority} / 类型: {case_type_export_label(case.case_type)}"
    child_topics = [
        _new_xmind_topic(f"前置条件: {case.preconditions or '无'}"),
        _new_xmind_topic(f"测试步骤: {case.steps}"),
        _new_xmind_topic(f"预期结果: {case.expected_result}"),
        _new_xmind_topic(f"属性: {attributes}"),
    ]
    return _new_xmind_topic(case.title, child_topics)


def export_cases_to_xmind(cases: list[ExportableFcCase], *, project_name: str) -> bytes:
    """Export functional test cases to XMind Zen compatible .xmind bytes."""
    if not cases:
        raise FcExportError("No test cases to export")

    modules: dict[str, list[ExportableFcCase]] = {}
    module_order: list[str] = []
    for case in cases:
        if case.module not in modules:
            modules[case.module] = []
            module_order.append(case.module)
        modules[case.module].append(case)

    module_topics = [
        _new_xmind_topic(module, [_build_case_topic(case) for case in modules[module]])
        for module in module_order
    ]
    root_topic = _new_xmind_topic(project_name, module_topics)

    content = [
        {
            "id": uuid.uuid4().hex,
            "class": "sheet",
            "title": "功能用例",
            "rootTopic": root_topic,
        }
    ]
    metadata = {
        "creator": {
            "name": "AITF",
            "version": "1.0.0",
        }
    }
    manifest = {
        "file-entries": {
            "content.json": {},
            "metadata.json": {},
        }
    }

    buffer = BytesIO()
    with zipfile.ZipFile(buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("content.json", json.dumps(content, ensure_ascii=False))
        archive.writestr("metadata.json", json.dumps(metadata, ensure_ascii=False))
        archive.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False))
    return buffer.getvalue()


__all__ = [
    "ExportableFcCase",
    "FcExportError",
    "case_to_excel_row",
    "case_type_export_label",
    "export_cases_to_excel",
    "export_cases_to_xmind",
    "safe_export_filename",
]
