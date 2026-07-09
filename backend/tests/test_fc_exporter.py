import json
import zipfile
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.services.fc_experience_importer import TEMPLATE_HEADERS
from app.services.fc_exporter import (
    ExportableFcCase,
    FcExportError,
    export_cases_to_excel,
    export_cases_to_xmind,
    safe_export_filename,
)

SAMPLE_CASES = [
    ExportableFcCase(
        case_no="FC-001",
        module="用户登录",
        title="正确密码登录",
        preconditions="用户已注册",
        steps="1. 打开登录页\n2. 输入账号密码",
        expected_result="登录成功",
        priority="P0",
        case_type="positive",
    ),
    ExportableFcCase(
        case_no="FC-002",
        module="用户登录",
        title="错误密码登录",
        preconditions=None,
        steps="1. 输入错误密码",
        expected_result="提示密码错误",
        priority="P1",
        case_type="negative",
    ),
]


def test_export_cases_to_excel_structure() -> None:
    content = export_cases_to_excel(SAMPLE_CASES, project_name="Demo Project")
    workbook = load_workbook(BytesIO(content))
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    assert headers == list(TEMPLATE_HEADERS)
    assert sheet.freeze_panes == "A2"
    assert sheet.max_row == 3
    assert sheet.cell(row=2, column=1).value == "FC-001"
    assert sheet.cell(row=2, column=8).value == "正向"
    assert sheet.cell(row=3, column=8).value == "异常"


def test_export_cases_to_excel_raises_when_empty() -> None:
    with pytest.raises(FcExportError):
        export_cases_to_excel([], project_name="Demo")


def test_export_cases_to_xmind_structure() -> None:
    content = export_cases_to_xmind(SAMPLE_CASES, project_name="Demo Project")
    with zipfile.ZipFile(BytesIO(content)) as archive:
        names = set(archive.namelist())
        assert "content.json" in names
        assert "metadata.json" in names
        payload = json.loads(archive.read("content.json").decode("utf-8"))

    assert isinstance(payload, list)
    sheet = payload[0]
    assert sheet["title"] == "功能用例"
    root = sheet["rootTopic"]
    assert root["title"] == "Demo Project"
    modules = root["children"]["attached"]
    assert len(modules) == 1
    assert modules[0]["title"] == "用户登录"
    case_topics = modules[0]["children"]["attached"]
    assert len(case_topics) == 2
    assert case_topics[0]["title"] == "正确密码登录"
    detail_titles = [item["title"] for item in case_topics[0]["children"]["attached"]]
    assert any(title.startswith("测试步骤:") for title in detail_titles)
    assert any("优先级: P0" in title for title in detail_titles)


def test_export_cases_to_xmind_raises_when_empty() -> None:
    with pytest.raises(FcExportError):
        export_cases_to_xmind([], project_name="Demo")


def test_safe_export_filename_ascii_only_for_chinese_name() -> None:
    assert safe_export_filename("测试项目", ".xlsx") == "fc-export.xlsx"
    assert safe_export_filename("Export Project", ".xlsx") == "Export_Project.xlsx"
